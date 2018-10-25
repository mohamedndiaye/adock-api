import datetime

from django.core.management.base import LabelCommand
from django.db import transaction
from openpyxl import load_workbook

from adock.carriers import models as carriers_models


class Command(LabelCommand):
    help = "Import labeled carriers from XLSX file"
    label = "XLSX file"

    HEADERS = (
        "entreprise",
        "entreprise incluse",
        "departement",
        "region",
        "siren",
        "millesime",
        "debut",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.counters = {"siren_not_found": 0, "county_not_found": 0, "found": 0}

    def import_objectif_co2(self, reader):
        # raw_row are tuples (read only)
        for i, raw_row in enumerate(reader):
            # Skip header (2 lines) and trailing empty lines
            if i < 2 or raw_row[0].value is None:
                continue

            # Cleanup
            row = {}
            for cell_index, header in enumerate(self.HEADERS):
                cell = raw_row[cell_index]

                if cell.data_type == "s":
                    row[header] = cell.value.replace("\n", " ").strip()
                else:
                    row[header] = cell.value

                # End date is start date + 3 years

            if isinstance(row["siren"], str):
                # Not always a string...
                row["siren"] = row["siren"].replace(" ", "")

            carriers = carriers_models.Carrier.objects.filter(
                siret__startswith=row["siren"]
            )
            if not carriers:
                self.counters["siren_not_found"] += 1
                self.stdout.write(
                    self.style.WARNING(
                        "Row %d: SIREN %s of %s not found."
                        % (i + 1, row["siren"], row["entreprise"])
                    )
                )
                continue

            carriers = carriers.filter(departement=row["departement"])
            if not carriers:
                self.counters["county_not_found"] += 1
                self.stdout.write(
                    self.style.WARNING(
                        "Row %d: SIREN %s and county %s of %s not found."
                        % (i + 1, row["siren"], row["departement"], row["entreprise"])
                    )
                )
                continue

            date_begin = datetime.datetime.date(row["debut"])
            carriers.update(
                objectif_co2=carriers_models.OBJECTIF_CO2_LABELLED,
                objectif_co2_begin=date_begin,
                # Same computing as in spreadsheet
                objectif_co2_end=date_begin.replace(year=date_begin.year + 3),
            )
            self.counters["found"] += 1
            self.stdout.write(
                "Row %d: %s"
                % (i + 1, [(t.raison_sociale, t.departement) for t in carriers])
            )

    def handle_label(self, xlsx_file, **options):  # pylint: disable=W0221
        self.stdout.write("Import XLSX file of labelled carriers: %s" % xlsx_file)

        wb = load_workbook(filename=xlsx_file, read_only=True)
        reader = wb.active.iter_rows()

        with transaction.atomic():
            # Reset
            carriers_models.Carrier.objects.update(
                objectif_co2="", objectif_co2_begin=None, objectif_co2_end=None
            )
            # Set
            self.import_objectif_co2(reader)

        self.stdout.write(self.style.SUCCESS("%s" % self.counters))
