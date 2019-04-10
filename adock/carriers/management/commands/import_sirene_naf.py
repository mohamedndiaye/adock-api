from django.core.management.base import LabelCommand
from django.db import connection
from django.db.utils import DataError
from openpyxl import load_workbook


class Command(LabelCommand):
    """
    Import XLS file from https://www.insee.fr/fr/information/2120875
    https://www.insee.fr/fr/statistiques/fichier/2120875/int_courts_naf_rev_2.xls
    The file should be converted to XLSX
    """

    help = "Import NAF labels from XLSX file"
    label = "XLSX file"

    HEADERS = ("_ligne", "code", "intitulé", "_intitulé 65", "_initulé 40")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def import_naf(self, cursor, reader):
        counter = 0
        for i, row in enumerate(reader):
            # Skip header (1 line) and trailing empty lines
            code = row[1].value
            if i < 1 or code is None or len(code) < 5 or code.startswith("SECTION"):
                continue

            code = code.strip()
            label = row[2].value.strip()
            try:
                cursor.execute(
                    "insert into sirene_naf (code, label) values (%s, %s)",
                    [code, label],
                )
            except DataError as e:
                self.stderr.write(self.style.ERROR("Error ligne %d") % i)
                self.stderr.write(self.style.ERROR("%s") % e)
            counter += 1

        return counter

    def handle_label(self, xlsx_file, **options):  # pylint: disable=W0221
        self.stdout.write("Import XLSX file of NAF labels: %s" % xlsx_file)

        wb = load_workbook(filename=xlsx_file, read_only=True)
        reader = wb.active.iter_rows()

        with connection.cursor() as cursor:
            cursor.execute("begin")
            cursor.execute("drop table if exists sirene_naf")
            cursor.execute("create table sirene_naf (code char(6), label char(256))")

            counter = self.import_naf(cursor, reader)

            cursor.execute("commit")

        self.stdout.write(self.style.SUCCESS("NAF codes imported: %s" % counter))
